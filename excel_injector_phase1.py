# -*- coding: utf-8 -*-
"""
Modul: excel_injector_fase1.py
Tanggung Jawab: Membuka template, menyuntikkan metadata administrasi global, 
               membangun topologi kaku sekuensial FAT & Tiang, serta melakukan
               pembersihan draf awal secara aman (Fail-Safe & Anti-Corrupt).
Arsitektur: Core Back-End Engine (Pass 1).
Developed by: An_
"""

import openpyxl
import io
import re
from system_config import RED_TAB_COLOR, WHITELIST_GLOBAL_SHEETS, BLACKLIST_CLUSTER, BLACKLIST_SUBFEEDER, METADATA_TOKEN_MAP

def is_red_tab(ws) -> bool:
    """
    Sensor Proteksi: Memeriksa apakah sheet memiliki warna tab MERAH (#FF0000).
    Mendukung pembacaan format Hex dengan/tanpa Alpha Channel dari openpyxl.
    """
    if ws.sheet_properties and ws.sheet_properties.tabColor and ws.sheet_properties.tabColor.rgb:
        color_str = str(ws.sheet_properties.tabColor.rgb)
        if RED_TAB_COLOR in color_str:
            return True
    return False

def inject_global_metadata(wb, metadata: dict):
    """
    Menyisipkan data administrasi umum (statis global) ke seluruh sheet 
    yang tidak dilindungi tab merah atau tidak masuk daftar blacklist.
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if is_red_tab(ws):
            continue  # Bypass Total absolut
            
        # Pindai seluruh sel aktif untuk menimpa token administrasi
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_str = cell.value
                    changed = False
                    for key, token in METADATA_TOKEN_MAP.items():
                        if token in cell_str and key in metadata:
                            cell_str = cell_str.replace(token, str(metadata[key]))
                            changed = True
                    if changed:
                        cell.value = cell_str

def fill_rigid_placeholders_column_major(wb, sheet_prefix: str, target_token: str, items_list: list, is_ba_splitter: bool = False):
    """
    SISTEM ALIRAN KAKU (COLUMN-MAJOR ORDER):
    Memindai token target pada kelompok sheet berawalan prefix (Cth: 'BA Splitter FAT_').
    Mengurutkan posisi secara Vertikal (Kolom dulu, baru baris), lalu mengisi sekuensial.
    Leftover slot otomatis dibersihkan secara aman (Border-Safe).
    """
    # Ambil dan urutkan sheet cadangan bawaan template yang sesuai prefix
    matched_sheets = sorted([s for s in wb.sheetnames if s.startswith(sheet_prefix)])
    
    item_index = 0
    used_sheets = set()

    for s_name in matched_sheets:
        ws = wb[s_name]
        if is_red_tab(ws):
            continue
            
        # Cari semua koordinat koordinat token di dalam sheet ini
        token_cells = []
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val and isinstance(cell_val, str) and target_token in cell_val:
                    token_cells.append((col, row, col, row)) # Simpan koordinat (Col, Row)
                    
        # SORTING KAKU: Column-Major Order (Kolom terkecil dulu, baru baris menurun ke bawah)
        token_cells.sort(key=lambda x: (x[0], x[1]))
        
        # Mulai mengalirkan data ke dalam kotak kaku
        for col, row, _, _ in token_cells:
            cell = ws.cell(row=row, column=col)
            if item_index < len(items_list):
                # Ganti token dengan data riil lapangan
                cell.value = str(cell.value).replace(target_token, items_list[item_index])
                item_index += 1
                used_sheets.add(s_name)
            else:
                # DATA HABIS: Lakukan pembersihan teks sisa agar tidak merusak visual
                if is_ba_splitter:
                    # Aturan Khusus BA Splitter: Bersihkan teks string 1 baris penuh dalam tabel
                    for c_idx in range(1, ws.max_column + 1):
                        # openpyxl aman menimpa sel kosong dengan "" tanpa merusak border
                        c_val = ws.cell(row=row, column=c_idx).value
                        if c_val and not any(tag in str(c_val) for tag in ["[DISTANCE]", "[1310]", "[1550]", "[INPUT_OPM]"]):
                            ws.cell(row=row, column=c_idx).value = ""
                else:
                    # Kasus Umum: Cukup kosongkan teks token itu sendiri
                    cell.value = str(cell.value).replace(target_token, "")

    return matched_sheets, used_sheets

def inject_excel_fase1(template_stream, metadata: dict, parsed_fat: list, parsed_poles: list, mode: str = "cluster"):
    """
    Fungsi Eksekutor Utama Pipeline Pass-1 (Fase 1).
    Merakit kerangka dokumen administratif kaku dan memotong lembar sisa.
    """
    # Memuat template perawan murni ke dalam memori RAM
    wb = openpyxl.load_workbook(template_stream)
    
    # 1. Suntik Seluruh Metadata Administrasi Global
    inject_global_metadata(wb, metadata)
    
    # 2. Penanganan Sektor Kaku FAT Individu & Judul Sheet Sekuensial
    fat_sheets = sorted([s for s in wb.sheetnames if s.startswith("FAT_")])
    used_fat_sheets = set()
    
    if mode == "cluster":
        # Jalur Cluster: Rename FAT_001 menjadi FAT A01, FAT_002 menjadi FAT A02, dst.
        for idx, fat_name in enumerate(parsed_fat):
            if idx < len(fat_sheets):
                orig_sheet = fat_sheets[idx]
                ws = wb[orig_sheet]
                ws.title = f"FAT {fat_name}"  # Nomenklatur resmi
                used_fat_sheets.add(ws.title)
                
                # Suntik token identitas internal sheet FAT
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and "[INPUT_FAT_NAME]" in str(cell.value):
                            cell.value = str(cell.value).replace("[INPUT_FAT_NAME]", f"FAT {fat_name}")
    else:
        # Jalur Subfeeder: Hanya butuh 1 sheet tunggal bernama murni "FAT"
        if fat_sheets:
            ws = wb[fat_sheets[0]]
            ws.title = "FAT"
            used_fat_sheets.add("FAT")
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and "[INPUT_FAT_NAME]" in str(cell.value):
                        cell.value = str(cell.value).replace("[INPUT_FAT_NAME]", "SUBFEEDER CORE")

    # 3. Penanganan Sektor Kaku Tiang (POLE Jaringan)
    pole_sheets = sorted([s for s in wb.sheetnames if s.startswith("POLE_")])
    used_pole_sheets = set()
    
    for idx, pole_data in enumerate(parsed_poles):
        # Setiap item kelompok tiang hasil parser mengisi 1 sheet POLE sekuensial bawaan template
        if idx < len(pole_sheets):
            orig_sheet = pole_sheets[idx]
            ws = wb[orig_sheet]
            ws.title = pole_data["title"]  # Rename menjadi 'Pole Erection 73' atau 'Pole Erection EXT 74'
            used_pole_sheets.add(ws.title)
            
            # Suntik spesifikasi deskripsi dan ejaan meter/inch tiang ke dalam placeholder kaku
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if "[INPUT_POLE_DESC]" in cell.value:
                            # Menghasilkan sekuensial kode komponen lokal formal (Cth: NEW POLE P001)
                            cell.value = cell.value.replace("[INPUT_POLE_DESC]", f"{pole_data['type']} {pole_data['title'].replace('Pole Erection ', '')}")
                        if "[INPUT_POLE_SIZE]" in cell.value:
                            cell.value = cell.value.replace("[INPUT_POLE_SIZE]", pole_data["size_clean"])

    # 4. Alirkan Topologi FAT ke Seluruh Berkas Rekap Pengukuran (Column-Major Flow)
    _, used_ba_splitters = fill_rigid_placeholders_column_major(wb, "BA Splitter FAT_", "[INPUT_FAT_NAME]", parsed_fat, is_ba_splitter=True)
    _, used_otdr_summaries = fill_rigid_placeholders_column_major(wb, "OTDR Sumary (FDT-FAT)_", "[INPUT_FAT_NAME]", parsed_fat, is_ba_splitter=False)
    _, used_opm_distributions = fill_rigid_placeholders_column_major(wb, "OPM_DISTRIBUTION_", "[INPUT_FAT_NAME]", parsed_fat, is_ba_splitter=False)

    # 5. GARBAGE COLLECTION: Pemusnahan Massal Lembar Sisa
    sheets_to_remove = []
    
    for s_name in wb.sheetnames:
        ws = wb[s_name]
        if is_red_tab(ws):
            continue  # Dilindungi mutlak oleh Tab Merah
            
        name_lower = s_name.lower().strip()
        
        # A. Hapus sheet master bernomor bawaan template yang tidak terpakai oleh kuotas data user
        if s_name.startswith("FAT_") or s_name.startswith("POLE_"):
            sheets_to_remove.append(s_name)
            continue
        if s_name.startswith("BA Splitter FAT_") and s_name not in used_ba_splitters:
            sheets_to_remove.append(s_name)
            continue
        if s_name.startswith("OTDR Sumary (FDT-FAT)_") and s_name not in used_otdr_summaries:
            sheets_to_remove.append(s_name)
            continue
        if s_name.startswith("OPM_DISTRIBUTION_") and s_name not in used_opm_distributions:
            sheets_to_remove.append(s_name)
            continue
            
        # B. Pemusnahan berbasis Aturan Mode Bisnis (Cluster vs Subfeeder) dari system_config
        if mode == "cluster":
            for blacklist in BLACKLIST_CLUSTER:
                if blacklist in name_lower:
                    sheets_to_remove.append(s_name)
                    break
        elif mode == "subfeeder":
            for blacklist in BLACKLIST_SUBFEEDER:
                if blacklist in name_lower:
                    sheets_to_remove.append(s_name)
                    break

    # Eksekusi penghapusan sheet dari workbook secara bersih
    for s_name in set(sheets_to_remove):
        if s_name in wb.sheetnames:
            wb.remove(wb[s_name])
            
    # Kembalikan file draf hidup berupa Bytes Stream ke dalam RAM memori
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream
