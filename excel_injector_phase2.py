# -*- coding: utf-8 -*-
"""
Modul: excel_injector_fase2.py
Tanggung Jawab: Menyuntikkan data angka teknis hasil ukur lapangan (OPM & OTDR) 
               ke dalam draf Fase 1 secara kaku, sekuensial, dan Column-Major 
               tanpa merusak borders XML (Anti-Corrupt & Safe Cleanup).
Arsitektur: Core Back-End Engine (Pass 2).
Developed by: An_
"""

import openpyxl
import io
from system_config import RED_TAB_COLOR

def check_red_tab_protection(ws) -> bool:
    """
    Sensor Proteksi: Memeriksa apakah sheet memiliki warna tab MERAH (#FF0000).
    Sheet bertab merah dilewati mutlak dari penyuntingan angka.
    """
    if ws.sheet_properties and ws.sheet_properties.tabColor and ws.sheet_properties.tabColor.rgb:
        if RED_TAB_COLOR in str(ws.sheet_properties.tabColor.rgb):
            return True
    return False

def inject_modul_2a_splitter_fdt(wb, df_splitter):
    """
    LOGIKA REKAP KAKU SPLITTER FDT (SIDE-BY-SIDE COLUMN MAJOR):
    Mengisi data dari Sisi Kiri Atas kebawah (Slot 1-2), lalu Sisi Kanan Atas kebawah (Slot 3-4).
    Token [OPM_AFTER] vertikal diisi berurutan berdasarkan indeks baris port.
    """
    splitter_sheets = sorted([s for s in wb.sheetnames if s.startswith("BA Splitter FDT") or s.startswith("OPM_")])
    if not splitter_sheets or df_splitter is None or df_splitter.empty:
        return

    records = df_splitter.to_dict(orient="records")
    record_idx = 0
    
    for s_name in splitter_sheets:
        ws = wb[s_name]
        if check_red_tab_protection(ws):
            continue
            
        # Peta koordinat kaku untuk token [SPL_ID] di dalam template
        # Mengelompokkan koordinat jangkar berdasarkan kolom fisik (Kiri vs Kanan)
        jankars = []
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=r, column=c).value == "[SPL_ID]":
                    jankars.append((c, r))
                    
        # Urutkan Column-Major: Sisi Kiri dulu (kolom kecil), baru Sisi Kanan (kolom besar)
        jankars.sort(key=lambda x: (x[0], x[1]))
        
        for c_jan, r_jan in jankars:
            if record_idx >= len(records):
                # Data habis: Bersihkan sisa token jangkar di template agar kosong bersih
                ws.cell(row=r_jan, column=c_jan).value = ""
                # Cari token [SPL_SN] dan [OPM_BEFORE] di area sekitar blok tersebut untuk dikosongkan
                for r_sweep in range(r_jan, r_jan + 12):
                    for c_sweep in range(max(1, c_jan-2), min(ws.max_column+1, c_jan+5)):
                        val = ws.cell(row=r_sweep, column=c_sweep).value
                        if val in ["[SPL_SN]", "[OPM_BEFORE]", "[OPM_AFTER]"]:
                            ws.cell(row=r_sweep, column=c_sweep).value = ""
                continue
                
            data = records[record_idx]
            record_idx += 1
            
            # Mulai isi blok kaku Splitter
            ws.cell(row=r_jan, column=c_jan).value = str(data.get("Splitter ID", ""))
            
            # Pindai ke bawah untuk mengisi SN, Before, dan 8 Port After sekuensial baris
            port_counter = 1
            for r_scan in range(r_jan, r_jan + 15):
                for c_scan in range(max(1, c_jan-2), min(ws.max_column+1, c_jan+5)):
                    cell = ws.cell(row=r_scan, column=c_scan)
                    if cell.value == "[SPL_SN]":
                        cell.value = str(data.get("Splitter SN", "NO SN"))
                    elif cell.value == "[OPM_BEFORE]":
                        cell.value = str(data.get("OPM Before (dBm)", ""))
                    elif cell.value == "[OPM_AFTER]":
                        p_val = data.get(f"Port {port_counter}", "")
                        cell.value = str(p_val) if p_val else "" # Kosong murni jika di UI blank
                        port_counter += 1

def inject_modul_2b_opm_distribution(wb, session_state_keys):
    """
    LOGIKA KAKU OPM DISTRIBUSI (SIDE-BY-SIDE 5 KOLOM BERJEJER KE KANAN):
    Mengisi data FAT dari Sisi Kiri Paling Atas, bergeser ke Kolom Kanan di sebelahnya.
    Setiap FAT diisi tegak lurus menurun untuk 8 baris port [INPUT_OPM].
    """
    dist_sheets = sorted([s for s in wb.sheetnames if s.startswith("OPM_DISTRIBUTION_")])
    
    # Kumpulkan seluruh baris data dari semua line grid editor di Web UI
    all_fat_records = []
    for key in session_state_keys:
        if key.startswith("grid_opm_dist_line_") and key in st.session_state:
            all_fat_records.extend(st.session_state[key].to_dict(orient="records"))
            
    if not dist_sheets or not all_fat_records:
        return

    fat_idx = 0
    for s_name in dist_sheets:
        ws = wb[s_name]
        if check_red_tab_protection(ws):
            continue
            
        # Cari semua sel yang berisi token [INPUT_FAT_NAME] hasil Pass-1
        f_jankars = []
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                # Di Fase 1, token ini sudah diubah menjadi nama FAT riil (Cth: FAT A01)
                if val and str(val).startswith("FAT "):
                    f_jankars.append((c, r, val))
                    
        # Urutkan Column-Major: Sisi Kiri ke Kanan (Kolom dulu, baru baris)
        f_jankars.sort(key=lambda x: (x[0], x[1]))
        
        for c_jan, r_jan, fat_title in f_jankars:
            # Cari record data UI yang namanya cocok dengan jangkar di sheet
            match_data = next((item for item in all_fat_records if item.get("FAT Name (Locked)") == fat_title), None)
            
            # Pindai secara vertikal ke bawah untuk mengisi 8 baris kaku token [INPUT_OPM]
            port_idx = 1
            for r_fill in range(r_jan + 1, r_jan + 10):
                for c_fill in range(max(1, c_jan-1), min(ws.max_column+1, c_jan+3)):
                    cell = ws.cell(row=r_fill, column=c_fill)
                    if cell.value == "[INPUT_OPM]":
                        if match_data:
                            p_val = match_data.get(f"Port {port_idx}", "")
                            cell.value = str(p_val) if p_val else ""
                        else:
                            cell.value = "" # Kosongkan jika tidak ada data lapangan
                        port_idx += 1

def inject_modul_2c_otdr_cluster(wb, df_otdr_cluster):
    """
    LOGIKA LOMPAT BARIS ZIG-ZAG OTDR CLUSTER:
    Mengisi Distance dan 1310 nm di baris ganjil yang sejajar dengan nama FAT,
    lalu melompat 1 baris ke bawahnya untuk mengisi token [1550] di baris genap.
    """
    otdr_sheets = sorted([s for s in wb.sheetnames if s.startswith("OTDR Sumary (FDT-FAT)_")])
    if not otdr_sheets or df_otdr_cluster is None or df_otdr_cluster.empty:
        return

    records = df_otdr_cluster.to_dict(orient="records")
    
    for s_name in otdr_sheets:
        ws = wb[s_name]
        if check_red_tab_protection(ws):
            continue
            
        # Pindai koordinat seluruh FAT di dalam tabel
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and str(val).startswith("FAT "):
                    match_data = next((item for item in records if item.get("FAT Name (Locked)") == val), None)
                    
                    # Isi sel ganjil (Distance dan 1310)
                    for c_sweep in range(c + 1, min(ws.max_column + 1, c + 5)):
                        cell_ganjil = ws.cell(row=r, column=c_sweep)
                        if cell_ganjil.value == "[DISTANCE]":
                            cell_ganjil.value = str(match_data.get("Distance (Km)", "")) if match_data else ""
                        elif cell_ganjil.value == "[1310]":
                            cell_ganjil.value = str(match_data.get("Loss 1310 nm (dB)", "")) if match_data else ""
                            
                    # LOGIKA LOMPAT KE BAWAH (Baris Genap untuk 1550 nm)
                    if r + 1 <= ws.max_row:
                        for c_sweep_genap in range(c + 1, min(ws.max_column + 1, c + 5)):
                            cell_genap = ws.cell(row=r + 1, column=c_sweep_genap)
                            if cell_genap.value == "[1550]":
                                cell_genap.value = str(match_data.get("Loss 1550 nm (dB)", "")) if match_data else ""

def inject_modul_2d_otdr_subfeeder(wb, df_otdr_subfeeder):
    """
    LOGIKA DATA SPLITTING SUBFEEDER:
    Mendistribusikan parameter jarak dan loss backbone utama ke sheet gelombang terpisah.
    - Sheet 1: Diubah judul [WAVE_LENGHT] -> 1310 nm, Tabel core diisi Loss 1310.
    - Sheet 2: Diubah judul [WAVE_LENGHT] -> 1550 nm, Tabel core diisi Loss 1550.
    """
    wave_sheets = sorted([s for s in wb.sheetnames if s.startswith("OTDR Summary (WAVE)_")])
    if len(wave_sheets) < 2 or df_otdr_subfeeder is None or df_otdr_subfeeder.empty:
        return

    records = df_otdr_subfeeder.to_dict(orient="records")

    # --- PART A: EKSEKUSI SHEET GELOMBANG 1310 NM ---
    ws_1310 = wb[wave_sheets[0]]
    if not check_red_tab_protection(ws_1310):
        # Ubah token judul
        for r in range(1, 5):
            for c in range(1, ws_1310.max_column + 1):
                if ws_1310.cell(row=r, column=c).value == "[WAVE_LENGHT]":
                    ws_1310.cell(row=r, column=c).value = "1310 nm"
                    
        # Isi tabel core kaku
        for r in range(5, ws_1310.max_row + 1):
            for c in range(1, ws_1310.max_column + 1):
                if ws_1310.cell(row=r, column=c).value == "[DISTANCE_SF]":
                    # Ambil indeks nomor core dari struktur tabel kaku
                    try:
                        core_num = int(ws_1310.cell(row=r, column=c-3).value)
                        if core_num <= len(records):
                            ws_1310.cell(row=r, column=c).value = str(records[core_num-1].get("Distance (Km)", ""))
                            ws_1310.cell(row=r, column=c+1).value = str(records[core_num-1].get("Loss 1310 nm (dB)", ""))
                        else:
                            ws_1310.cell(row=r, column=c).value = ""
                            ws_1310.cell(row=r, column=c+1).value = ""
                    except:
                        pass

    # --- PART B: EKSEKUSI SHEET GELOMBANG 1550 NM ---
    ws_1550 = wb[wave_sheets[1]]
    if not check_red_tab_protection(ws_1550):
        # Ubah token judul
        for r in range(1, 5):
            for c in range(1, ws_1550.max_column + 1):
                if ws_1550.cell(row=r, column=c).value == "[WAVE_LENGHT]":
                    ws_1550.cell(row=r, column=c).value = "1550 nm"
                    
        # Isi tabel core kaku
        for r in range(5, ws_1550.max_row + 1):
            for c in range(1, ws_1550.max_column + 1):
                if ws_1550.cell(row=r, column=c).value == "[DISTANCE_SF]":
                    try:
                        core_num = int(ws_1550.cell(row=r, column=c-3).value)
                        if core_num <= len(records):
                            ws_1550.cell(row=r, column=c).value = str(records[core_num-1].get("Distance (Km)", ""))
                            ws_1550.cell(row=r, column=c+1).value = str(records[core_num-1].get("Loss 1550 nm (dB)", ""))
                        else:
                            ws_1550.cell(row=r, column=c).value = ""
                            ws_1550.cell(row=r, column=c+1).value = ""
                    except:
                        pass

def inject_excel_fase2(draf_stream, mode: str = "cluster"):
    """
    Fungsi Eksecutor Utama Pipeline Pass-2 (Fase 2).
    Menangkap draf hidup dari Fase 1 di RAM, menyuntikkan seluruh parameter angka,
    dan mengembalikan objek file final utuh siap unduh.
    """
    import streamlit as st # Impor lokal aman di dalam fungsi
    wb = openpyxl.load_workbook(draf_stream)

    # 1. Eksekusi Suntik Angka Modul 2A (Splitter FDT & Feeder OPM)
    df_splitter = st.session_state.get("grid_splitter_data", None)
    inject_modul_2a_splitter_fdt(wb, df_splitter)

    # 2. Eksekusi Suntik Angka Modul 2B (OPM Distribusi Massal)
    inject_modul_2b_opm_distribution(wb, list(st.session_state.keys()))

    # 3. Eksekusi Suntik Angka Sektor OTDR Berdasarkan Mode Bisnis
    if mode == "cluster":
        df_otdr_cluster = st.session_state.get("grid_otdr_cluster_data", None)
        inject_modul_2c_otdr_cluster(wb, df_otdr_cluster)
    else:
        df_otdr_subfeeder = st.session_state.get("grid_otdr_subfeeder_data", None)
        inject_modul_2d_otdr_subfeeder(wb, df_otdr_subfeeder)

    # 4. Final Pembersihan Sisa: Pastikan tidak ada token teknis mentah tertinggal di berkas final
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if check_red_tab_protection(ws):
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and any(tag in str(cell.value) for tag in ["[DISTANCE]", "[1310]", "[1550]", "[INPUT_OPM]", "[OPM_BEFORE]", "[OPM_AFTER]", "[SPL_ID]", "[SPL_SN]"]):
                    cell.value = ""

    # Bungkus berkas matang terakhir ke dalam biner stream RAM
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output
